import os
import getopt
import sys
from pathlib import PurePath
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Color, PatternFill, Font, Border

#from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
#from openpyxl.utils import get_column_letter


RULE_PKG = "For ivsc_pkg_{sensor}_0_*.bin, 1) its size 500KB"
RULE_SKUCFG = "For ivsc_skucfg_{sensor}_0_*.bin, size between 4KB to 6KB"
RULE_VCFG_1 ="For ivsc_vcfg_{sensor}_1_*.bin, size 32-byte and data 0-0xF or 0x20"
RULE_VCFG_2 = "For ivsc_vcfg_{sensor}_2_*.bin, type text with keyword place-holder"


def check_pkg(full_fname):
    size = os.path.getsize(full_fname)
    if size > 500*1024:
        return True, ''
    return False, 'file size < 5KB'


def check_skucfg(full_fname):
    size = os.path.getsize(full_fname)
    if 4000 < size < 6000:
        return True, ''
    return False, 'file size > 6KB or <4 KB'


def check_vcfg_1(full_fname):
    with open(full_fname, 'rb') as f:
        buffer = f.read()
        #assert(len(buffer) == 32)
        if len(buffer) != 32:
            return False, 'file size != 32'
        for i in buffer:
            # binary data must have value 0-15 (0xF) or 32 (0x20) 
            if i > 16 and i != 0x20: 
                '''
                cnt = 0
                print('Error: unexpected val %02X' % (i), 'from', full_fname)
                print('DUMP: ', end='')
                for k in buffer: 
                    cnt += 1
                    if cnt % 4 == 0:
                        cnt = 0
                        print(" %02X" % (k), end='')
                    else: 
                        print("%02X" % (k), end='')
                print()
                '''
                return False, 'data value > 15 and != 32'
        #print('File', file, 'has', len(buffer), 'bytes... ok!')

    return True, ''


def check_cfg_placeholder(full_fname):
    with open(full_fname) as f:
        buffer = f.read()

    if 'place-holder' in buffer:
        return True, ''
            
    return False, 'no place-holder keyward found' 


def workbook_header(sheet, root):

    sheet.append((root, '','',''))
    sheet.append(('','','',''))
    sheet.append(('','File name', 'Status', 'Reason'))
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['B3'].font = Font(bold=True)
    sheet['C3'].font = Font(bold=True)
    sheet['D3'].font = Font(bold=True)
    sheet.column_dimensions['A'].height = 40
    #Fill = PatternFill("solid", fgColor="DDDDDD")
    #sheet['A1'].fill = PatternFill("solid", start_color="5cb800") 
    return



def walk_dir(folder):
    
    # Open log and excel
    fd  = open('FW_CHK.log', 'w')
    fd2 = open('FW_CHK_ERR.log', 'w')
    wb = openpyxl.Workbook()

    for root, dirs, files in os.walk(folder):

        # Skip files not interested
        if 'MUP' in root or os.path.basename(root) == 'firmware' or \
           os.path.basename(root) == 'IVSC' or os.path.basename(root) == 'Debug':
            continue
       
        # Only look for 'IVSC folder'
        if r'Drivers\IVSC' in root:

            path = os.path.dirname(root)     # c:\tmp\RPL\SIK_release_WW31\ivsc-3012\Drivers\IVSC
            sensor = os.path.basename(root)  # ovti9738

            # Skip these two folders 
            if sensor == 'Symbols' or sensor == 'Public':
                continue

            print(root)   
            fd.write('\n' + root + '\n')

            # Prepare a workbook
            sheet = wb.create_sheet(sensor)  
            workbook_header(sheet, root)

            for file in files:
                # File name: ivsc_pkg_ovti9738_0_a0.bin
                full_fname = PurePath(os.path.join(root,file))
                status = None
                reason = ''
                if file.startswith(f'ivsc_pkg_{sensor}_0'):
                    status, reason =  check_pkg(full_fname)
                    out = '%40s checked... %s' % (file, 'ok' if status else 'FAILED')

                elif file.startswith(f'ivsc_skucfg_{sensor}_0_1'):
                    status, reason =  check_skucfg(full_fname)
                    out = '%40s checked... %s' % (file, 'ok' if status else 'FAILED')


                elif file.startswith(f'ivsc_skucfg_{sensor}_0_2'):
                    status, reason = check_cfg_placeholder(full_fname)
                    out = '%40s checked... %s' % (file, 'ok' if status else 'FAILED')

                elif file.startswith(f'ivsc_vcfg_{sensor}_1'):
                    status, reason = check_vcfg_1(full_fname) 
                    out = '%40s checked... %s' % (file, 'ok' if status else 'FAILED')


                elif file.startswith(f'ivsc_vcfg_{sensor}_2'):
                    status, reason = check_cfg_placeholder(full_fname) 
                    out = '%40s checked... %s' % (file, 'ok' if status else 'FAILED')

                else:
                    reason = 'has NO RULE to check'
                    out = '%40s has NO RULE to check' % (file)

                print(out)
                fd.write(out + '\n')

                # For xlsx
                if status is True:
                    sheet.append(('', file, 'OK', ''))
                elif status is False:
                    sheet.append(('', file, 'FAILED', reason)) 
                    row = sheet._current_row
                    pos = 'C{}'.format(row)
                    sheet[pos].fill = PatternFill("solid", start_color="FF0000")

                    # Error log such as 'File FAILED no place-holder keyward found'
                    fd2.write(f'{full_fname}, FAILED {reason}\n')

            # Done with all sensors under this root/folder
            print()
            fd.write('\n')
            
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['D'].width = 60


    # Write, save and close
    wb.remove(wb['Sheet'])
    wb.save('FW_CHK.xlsx')
    fd.close()    
    fd2.close()
    
    print('Completed. Log files FW_CHK.log, FW_CHK_ERROR.log and FW_CHK.xlsx\n')
    return 

def main():
    opts, args = getopt.getopt(sys.argv[1:], "hd:", ["help", "dir"])

    folder = os.getcwd() 
    for name, val in opts:
        if name in ('-h', '?', '--help'):
            print('Usage\n     python3 {} [-h] [-d dir]\n'.format(sys.argv[0]))
            print('Example\n     python3 {}'.format(sys.argv[0]))
            print(r'     python3 {} -d "Y:\Karuna\RPL\SIK_release_WW31"'.format(sys.argv[0]))
            print(r'     python3 {} -d "SIK_release_WW29"'.format(sys.argv[0]))
            print()
            print(f'Inspection Rules')
            print(f'     {RULE_PKG}')
            print(f'     {RULE_SKUCFG}')
            print(f'     {RULE_VCFG_1}')
            print(f'     {RULE_VCFG_2}')
            print()
            sys.exit(0)
        if name in ('-d', '--dir'):
            folder = val


    walk_dir(folder)
    return
    
if __name__ == '__main__':
    main()

        
